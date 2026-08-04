"""Parse CSV / XLSX uploads into header + row dicts."""

from __future__ import annotations

import csv
import io
from typing import Any


class ParseError(ValueError):
    pass


def _normalize_header(value: Any) -> str:
    return str(value or "").strip()


def parse_csv_bytes(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ParseError("CSV has no header row.")
    headers = [_normalize_header(h) for h in reader.fieldnames if _normalize_header(h)]
    rows: list[dict[str, str]] = []
    for raw in reader:
        row = { _normalize_header(k): str(v or "").strip() for k, v in raw.items() if _normalize_header(k) }
        if any(row.values()):
            rows.append(row)
    return headers, rows


def parse_xlsx_bytes(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError("Excel support requires openpyxl.") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    # Prefer a sheet named "Students"; else first data sheet (skip Instructions).
    sheet = None
    for name in wb.sheetnames:
        if name.lower() == "students":
            sheet = wb[name]
            break
    if sheet is None:
        for name in wb.sheetnames:
            if name.lower() != "instructions":
                sheet = wb[name]
                break
    if sheet is None:
        raise ParseError("Workbook has no data sheet.")

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ParseError("Excel sheet is empty.") from exc

    headers = [_normalize_header(h) for h in header_row if _normalize_header(h)]
    if not headers:
        raise ParseError("Excel sheet has no header row.")

    rows: list[dict[str, str]] = []
    for values in rows_iter:
        if values is None:
            continue
        raw = {}
        for idx, header in enumerate(headers):
            cell = values[idx] if idx < len(values) else ""
            if cell is None:
                cell = ""
            raw[header] = str(cell).strip()
        if any(raw.values()):
            rows.append(raw)
    return headers, rows


def parse_upload(filename: str, content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return parse_xlsx_bytes(content)
    if lower.endswith(".csv") or lower.endswith(".txt"):
        return parse_csv_bytes(content)
    # Sniff: xlsx is a zip (PK…)
    if content[:2] == b"PK":
        return parse_xlsx_bytes(content)
    return parse_csv_bytes(content)


def build_csv_template() -> bytes:
    from apps.admissions.imports.columns import CANONICAL_KEYS, SAMPLE_ROWS

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CANONICAL_KEYS, extrasaction="ignore")
    writer.writeheader()
    for row in SAMPLE_ROWS:
        writer.writerow(row)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_xlsx_template() -> bytes:
    from openpyxl import Workbook
    from apps.admissions.imports.columns import CANONICAL_KEYS, INSTRUCTIONS, SAMPLE_ROWS

    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    for i, line in enumerate(INSTRUCTIONS.strip().split("\n"), start=1):
        instructions.cell(row=i, column=1, value=line)

    students = wb.create_sheet("Students")
    for col, key in enumerate(CANONICAL_KEYS, start=1):
        students.cell(row=1, column=col, value=key)
    for r_idx, row in enumerate(SAMPLE_ROWS, start=2):
        for c_idx, key in enumerate(CANONICAL_KEYS, start=1):
            students.cell(row=r_idx, column=c_idx, value=row.get(key, ""))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
